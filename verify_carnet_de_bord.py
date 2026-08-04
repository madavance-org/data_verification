#!/usr/bin/env python3
"""
Vérification de données — Carnet de bord
=========================================

Applique cinq des six dimensions du "Manuel de vérification de données" MadAvance
(Complétude, Promptitude, Validité, Unicité, Cohérence) à l'activité "Carnet de
bord" (suivi véhicules : trajets, carburant, lavage, entretien/maintenance,
renouvellement de documents administratifs).

La sixième dimension, Fiabilité, n'est PAS automatisée ici : elle consiste en un
rapprochement documentaire (facture/fiche physique <-> saisie mWater), qui reste
un contrôle manuel volontairement hors de portée d'un script (voir le manuel
ClickUp lié).

Le script télécharge le datagrid mWater "Carnet de bord" (déjà configuré dans le
portail, même mécanisme que verify_maintenance_preventive.py dans ce repo),
applique les règles, puis met à jour un log Excel cumulatif
(data_verification_carnet_de_bord_log.xlsx, fichier distinct du log Appel
maintenance préventive) qu'il dépose sur SharePoint via Microsoft Graph, et
envoie un email de confirmation.

Réutilise l'authentification mWater/Graph et la logique de log (Nouveau /
Toujours ouvert / Résolu) de verify_maintenance_preventive.py dans ce même repo.
"""

import base64
import csv
import difflib
import io
import math
import os
import sys
from collections import Counter, defaultdict
from datetime import datetime

import requests

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

MWATER_API_BASE = "https://api.mwater.co/v3"

# ID du datagrid mWater "Carnet de bord" (préconfiguré dans le portail, pas un secret)
DATAGRID_CARNET = "d30af8d9ab7b4bb3b0aae2adbcce622f"

# Fichier de log dédié à cette activité — distinct de data_verification_call_log.xlsx
# (celui d'Appel maintenance préventive), même s'il vit dans le même dossier SharePoint.
LOG_FILE_NAME = "data_verification_carnet_de_bord_log.xlsx"

LOG_HEADERS = [
    "Dimension", "Sous-dimension", "Response Code", "Véhicule",
    "Description", "Détails", "Statut", "Première détection",
    "Dernière détection", "Date de résolution",
]

# Dimensions couvertes par ce script (Fiabilité exclue -- voir docstring).
DIMENSIONS_AUTOMATISEES = ["Complétude", "Promptitude", "Validité", "Unicité", "Cohérence"]

TYPE_TRAJET = "Enregistrement trajet"
TYPE_CARBURANT = "Approvisionnement en Carburant"
TYPE_LAVAGE = "Lavage"
TYPE_ENTRETIEN = "Entretien et Maintenance"
TYPE_RENOUVELLEMENT = "Renouvellement des Documents Administratifs"

# Distance maximale plausible pour un seul trajet (compteur), en km. Sert à détecter les
# compteurs corrompus indépendamment du GPS (ex. plusieurs millions de km vus dans les données
# réelles). Arbitraire mais large -- à ajuster si des trajets légitimes longue distance existent.
COMPTEUR_TRAJET_MAX_KM = 2000

# Précision GPS par défaut (mètres) si le champ 'accuracy' est absent/nul pour un point --
# confirmé par Lanja : certains appareils utilisés ont une précision de l'ordre de 3 km.
GPS_ACCURACY_DEFAUT_M = 3000

# Marge multiplicative appliquée à la précision GPS rapportée, et plancher de tolérance (km),
# pour la comparaison distance GPS vs compteur. Arbitraires -- à affiner avec Lanja au fil de
# l'usage réel du script.
GPS_TOLERANCE_MARGE = 2.0
GPS_TOLERANCE_PLANCHER_KM = 5.0

# Seuil de similarité (0-1) au-delà duquel deux noms de conducteur différents sont
# signalés comme variante probable (faute de frappe) plutôt qu'ignorés. Arbitraire --
# à valider avec Lanja ; whitelister les vrais homonymes si des faux positifs apparaissent.
SIMILARITE_NOM_SEUIL = 0.82


def raise_for_status_verbose(response):
    """Comme dans verify_maintenance_preventive.py : loggue le corps complet en cas d'erreur HTTP."""
    if not response.ok:
        print(f"HTTP {response.status_code} sur {response.url}", file=sys.stderr)
        print(response.text[:2000], file=sys.stderr)
        response.raise_for_status()


# ---------------------------------------------------------------------------
# mWater : authentification et téléchargement du datagrid
# ---------------------------------------------------------------------------

def mwater_login(username, password):
    resp = requests.post(
        f"{MWATER_API_BASE}/clients",
        json={"username": username, "password": password},
        timeout=30,
    )
    raise_for_status_verbose(resp)
    client_id = resp.json().get("client")
    if not client_id:
        raise RuntimeError("Authentification mWater : champ 'client' absent de la réponse")
    return client_id


def download_datagrid(datagrid_id, client_id):
    """Télécharge un datagrid mWater et le retourne comme liste de dict (une par ligne)."""
    resp = requests.get(
        f"{MWATER_API_BASE}/datagrids/{datagrid_id}/download",
        params={"client": client_id, "share": "", "extraFilters": "[]", "format": "csv"},
        timeout=120,
    )
    raise_for_status_verbose(resp)
    text = resp.content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return list(reader)


# ---------------------------------------------------------------------------
# Utilitaires de parsing
# ---------------------------------------------------------------------------

def parse_dt(value):
    """Parse un horodatage mWater 'AAAA-MM-JJ HH:MM:SS' ou une date 'AAAA-MM-JJ'."""
    value = (value or "").strip()
    if not value:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    return None


def parse_float(value):
    value = (value or "").strip()
    if not value:
        return None
    try:
        return float(value)
    except ValueError:
        return None


def parse_latlng(value):
    """Parse une valeur de localisation mWater 'lat, lng'. None si absente/invalide."""
    value = (value or "").strip()
    if not value or "," not in value:
        return None
    try:
        lat_str, lng_str = value.split(",", 1)
        return float(lat_str.strip()), float(lng_str.strip())
    except ValueError:
        return None


def haversine_km(p1, p2):
    """Distance à vol d'oiseau entre deux points (lat, lng), en km."""
    lat1, lng1 = p1
    lat2, lng2 = p2
    r = 6371.0
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lng2 - lng1)
    a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2) ** 2
    return 2 * r * math.asin(math.sqrt(a))


# ---------------------------------------------------------------------------
# Anomalie : structure commune
# ---------------------------------------------------------------------------

class Anomaly:
    def __init__(self, dimension, subdimension, response_code, vehicule, description, details=""):
        self.dimension = dimension
        self.subdimension = subdimension
        self.response_code = response_code
        self.vehicule = vehicule
        self.description = description
        self.details = details


# ---------------------------------------------------------------------------
# Dimension : Complétude
# ---------------------------------------------------------------------------

def check_completude(rows):
    anomalies = []
    for r in rows:
        if r.get("Status") == "Draft":
            anomalies.append(Anomaly(
                "Complétude", "Brouillon non soumis",
                r.get("Response Code", ""), r.get("Véhicule > Immatriculaton", ""),
                "Brouillon jamais soumis",
                f"Drafted On: {r.get('Drafted On', '')}",
            ))
    return anomalies


# ---------------------------------------------------------------------------
# Dimension : Promptitude
# ---------------------------------------------------------------------------

def event_dates_for_row(r):
    """Renvoie les champs date propres à l'événement décrit par la ligne (pas les métadonnées
    de saisie Drafted On/Submitted On), selon le type d'intervention -- sert à vérifier que la
    date de l'événement correspond bien au mois de la saisie (voir check_promptitude_date_evenement)
    et à exclure du calcul Cohérence les lignes dont la date est manifestement fausse."""
    type_intervention = (r.get("Type d'intervention") or "").strip()
    if type_intervention == TYPE_TRAJET:
        return [("Heure de départ", r.get("Heure de départ")), ("Heure d'arrivée", r.get("Heure d'arrivée"))]
    return [("Date", r.get("Date"))]


def check_promptitude(rows):
    anomalies = []
    for r in rows:
        drafted = parse_dt(r.get("Drafted On"))
        submitted = parse_dt(r.get("Submitted On"))
        if drafted and submitted and submitted < drafted:
            anomalies.append(Anomaly(
                "Promptitude", "Chronologie Drafted/Submitted",
                r.get("Response Code", ""), r.get("Véhicule > Immatriculaton", ""),
                "Submitted On antérieur à Drafted On",
                f"Drafted On: {r.get('Drafted On', '')} / Submitted On: {r.get('Submitted On', '')}",
            ))
    anomalies += check_promptitude_date_evenement(rows)
    return anomalies


def check_promptitude_date_evenement(rows):
    """Une saisie faite un mois donné doit décrire un événement du même mois (même année et
    même mois que Drafted On) -- précisé par Lanja suite aux dates manifestement aberrantes
    trouvées dans les données réelles (1908, 2002, 2014, 2015...). Comparaison relative à
    Drafted On plutôt qu'à un seuil absolu (ex. 'avant 2020') : reste valide sans mise à jour
    au fil du temps, et n'exclut pas à tort les saisies rétroactives légitimes d'un mois sur
    l'autre -- seul un écart de mois complet est signalé."""
    anomalies = []
    for r in rows:
        drafted = parse_dt(r.get("Drafted On"))
        if not drafted:
            continue
        for label, raw_value in event_dates_for_row(r):
            event_dt = parse_dt(raw_value)
            if not event_dt:
                continue
            if (event_dt.year, event_dt.month) != (drafted.year, drafted.month):
                anomalies.append(Anomaly(
                    "Promptitude", "Date événement hors mois de saisie",
                    r.get("Response Code", ""), r.get("Véhicule > Immatriculaton", ""),
                    f"{label} ne correspond pas au mois de la saisie (Drafted On)",
                    f"{label}: {raw_value} / Drafted On: {r.get('Drafted On', '')}",
                ))
    return anomalies


# ---------------------------------------------------------------------------
# Dimension : Validité
# ---------------------------------------------------------------------------

def check_validite(rows):
    anomalies = []
    for r in rows:
        rc = r.get("Response Code", "")
        vehicule = r.get("Véhicule > Immatriculaton", "")
        type_intervention = (r.get("Type d'intervention") or "").strip()

        if type_intervention == TYPE_TRAJET:
            depart = parse_float(r.get("Compteur de départ"))
            arrivee = parse_float(r.get("Compteur d'arrivée"))
            if depart is not None and arrivee is not None and arrivee < depart:
                anomalies.append(Anomaly(
                    "Validité", "Distance négative", rc, vehicule,
                    "Compteur d'arrivée inférieur au compteur de départ",
                    f"Départ: {depart} / Arrivée: {arrivee}",
                ))

            # Accuracy absente = position potentiellement non fiable (ex. saisie manuelle du
            # point plutôt que capture GPS de l'appareil) -- signalé indépendamment de tout
            # calcul de distance, précisé par Lanja : "on ne sait pas avec précision si c'est
            # la vrai position" dès lors que ce champ est vide.
            p_depart = (r.get("Point de départ") or "").strip()
            if p_depart and not (r.get("Point de départ (accuracy)") or "").strip():
                anomalies.append(Anomaly(
                    "Validité", "Position GPS sans précision (accuracy)", rc, vehicule,
                    "Point de départ renseigné sans valeur d'accuracy -- fiabilité de la position incertaine",
                    f"Point de départ: {p_depart}",
                ))
            p_arrivee = (r.get("Point d'arrivée") or "").strip()
            if p_arrivee and not (r.get("Point d'arrivée (accuracy)") or "").strip():
                anomalies.append(Anomaly(
                    "Validité", "Position GPS sans précision (accuracy)", rc, vehicule,
                    "Point d'arrivée renseigné sans valeur d'accuracy -- fiabilité de la position incertaine",
                    f"Point d'arrivée: {p_arrivee}",
                ))

        elif type_intervention == TYPE_CARBURANT:
            compteur = parse_float(r.get("Compteur au moment de l'approvisionnement en carburant"))
            if compteur is not None and compteur <= 0:
                anomalies.append(Anomaly(
                    "Validité", "Compteur non positif", rc, vehicule,
                    "Compteur au moment de l'approvisionnement <= 0",
                    f"Valeur: {compteur}",
                ))
            litres = parse_float(r.get("Quantité du carburant injecté (Litre)"))
            if litres is not None and litres <= 0:
                anomalies.append(Anomaly(
                    "Validité", "Litres non positifs", rc, vehicule,
                    "Quantité de carburant <= 0",
                    f"Valeur: {litres}",
                ))
            prix = parse_float(r.get("Prix Unitaire"))
            if prix is not None and prix <= 0:
                anomalies.append(Anomaly(
                    "Validité", "Prix non positif", rc, vehicule,
                    "Prix unitaire du carburant <= 0",
                    f"Valeur: {prix}",
                ))

        elif type_intervention == TYPE_ENTRETIEN:
            compteur = parse_float(r.get("Compteur au moment de l'entretien/maintenance"))
            if compteur is not None and compteur <= 0:
                anomalies.append(Anomaly(
                    "Validité", "Compteur non positif", rc, vehicule,
                    "Compteur au moment de l'entretien/maintenance <= 0",
                    f"Valeur: {compteur}",
                ))
            cout = parse_float(r.get("Coût total de l'entretien et/ou maintenance"))
            if cout is not None and cout <= 0:
                anomalies.append(Anomaly(
                    "Validité", "Coût non positif", rc, vehicule,
                    "Coût total entretien/maintenance <= 0",
                    f"Valeur: {cout}",
                ))

        elif type_intervention == TYPE_LAVAGE:
            cout = parse_float(r.get("Coût service de lavage"))
            if cout is not None and cout <= 0:
                anomalies.append(Anomaly(
                    "Validité", "Coût non positif", rc, vehicule,
                    "Coût du service de lavage <= 0",
                    f"Valeur: {cout}",
                ))

        elif type_intervention == TYPE_RENOUVELLEMENT:
            montant = parse_float(r.get("Montant du renouvellement"))
            if montant is not None and montant <= 0:
                anomalies.append(Anomaly(
                    "Validité", "Montant non positif", rc, vehicule,
                    "Montant du renouvellement <= 0",
                    f"Valeur: {montant}",
                ))
            date_saisie = parse_dt(r.get("Date"))
            nouvelle_limite = parse_dt(r.get("Nouvelle date limite de validation"))
            if date_saisie and nouvelle_limite and nouvelle_limite <= date_saisie:
                anomalies.append(Anomaly(
                    "Validité", "Date limite non future", rc, vehicule,
                    "Nouvelle date limite de validité antérieure ou égale à la date de saisie",
                    f"Date: {r.get('Date', '')} / Nouvelle date limite: {r.get('Nouvelle date limite de validation', '')}",
                ))
    return anomalies


# ---------------------------------------------------------------------------
# Dimension : Unicité
# ---------------------------------------------------------------------------

def check_unicite(rows):
    """Doublon = deux saisies de trajet partageant le même Compteur de départ, pour le
    même véhicule (précisé par Lanja -- deux véhicules différents peuvent coïncidemment
    partager un relevé sans que ce soit une vraie anomalie)."""
    anomalies = []
    trajets = [r for r in rows if (r.get("Type d'intervention") or "").strip() == TYPE_TRAJET]

    counts = Counter(
        (r.get("Véhicule > Immatriculaton", ""), (r.get("Compteur de départ") or "").strip())
        for r in trajets
        if (r.get("Compteur de départ") or "").strip()
    )
    duplicated = {key for key, n in counts.items() if n > 1}

    for r in trajets:
        key = (r.get("Véhicule > Immatriculaton", ""), (r.get("Compteur de départ") or "").strip())
        if key in duplicated:
            anomalies.append(Anomaly(
                "Unicité", "Doublon compteur de départ",
                r.get("Response Code", ""), r.get("Véhicule > Immatriculaton", ""),
                "Deux saisies de trajet partagent le même compteur de départ pour ce véhicule",
                f"Compteur de départ: {r.get('Compteur de départ', '')}",
            ))
    return anomalies


# ---------------------------------------------------------------------------
# Dimension : Cohérence
# ---------------------------------------------------------------------------

def check_coherence_compteur_croissant(rows):
    """Pour chaque véhicule, le compteur de chaque nouvelle entrée (trajet départ/arrivée,
    carburant, entretien) doit être >= au dernier compteur connu, toutes interventions
    confondues et triées chronologiquement -- le kilométrage ne peut que croître."""
    anomalies = []

    # (timestamp, valeur_compteur, response_code, description_horodatage) par véhicule
    events_by_vehicule = defaultdict(list)

    for r in rows:
        vehicule = r.get("Véhicule > Immatriculaton", "")
        if not vehicule:
            continue
        type_intervention = (r.get("Type d'intervention") or "").strip()
        rc = r.get("Response Code", "")
        drafted = parse_dt(r.get("Drafted On"))

        def date_fiable(dt):
            """Exclut du tri chronologique les dates dont le mois ne correspond pas à celui de
            la saisie (Drafted On) -- déjà signalées par check_promptitude_date_evenement,
            sinon elles faussent en cascade tout l'historique du véhicule (voir données réelles :
            une date de 1908/2002/2014/2015 isolée corrompait des centaines de comparaisons)."""
            if not dt or not drafted:
                return dt is not None
            return (dt.year, dt.month) == (drafted.year, drafted.month)

        if type_intervention == TYPE_TRAJET:
            depart_t = parse_dt(r.get("Heure de départ"))
            depart_c = parse_float(r.get("Compteur de départ"))
            if depart_t and depart_c is not None and date_fiable(depart_t):
                events_by_vehicule[vehicule].append((depart_t, depart_c, rc, "Compteur de départ (trajet)"))
            arrivee_t = parse_dt(r.get("Heure d'arrivée"))
            arrivee_c = parse_float(r.get("Compteur d'arrivée"))
            if arrivee_t and arrivee_c is not None and date_fiable(arrivee_t):
                events_by_vehicule[vehicule].append((arrivee_t, arrivee_c, rc, "Compteur d'arrivée (trajet)"))
        elif type_intervention == TYPE_CARBURANT:
            t = parse_dt(r.get("Date"))
            c = parse_float(r.get("Compteur au moment de l'approvisionnement en carburant"))
            if t and c is not None and date_fiable(t):
                events_by_vehicule[vehicule].append((t, c, rc, "Compteur (carburant)"))
        elif type_intervention == TYPE_ENTRETIEN:
            t = parse_dt(r.get("Date"))
            c = parse_float(r.get("Compteur au moment de l'entretien/maintenance"))
            if t and c is not None and date_fiable(t):
                events_by_vehicule[vehicule].append((t, c, rc, "Compteur (entretien)"))

    # Pas de filtre par ampleur de l'écart (voir échange avec Lanja) : un petit écart répété
    # peut être le signe d'une manipulation volontaire (ex. sous-déclaration pour justifier du
    # carburant), pas seulement du bruit de lecture -- filtrer par magnitude masquerait
    # justement ce type de signal. On garde donc tous les écarts, et on ajoute le nombre total
    # d'occurrences pour ce véhicule dans les détails, pour aider un relecteur humain à
    # distinguer un cas isolé d'un motif répété.
    par_vehicule = defaultdict(list)
    for vehicule, events in events_by_vehicule.items():
        events.sort(key=lambda e: e[0])
        for prev, curr in zip(events, events[1:]):
            prev_t, prev_c, prev_rc, prev_desc = prev
            curr_t, curr_c, curr_rc, curr_desc = curr
            if curr_c < prev_c:
                par_vehicule[vehicule].append((curr_rc, prev_desc, prev_t, prev_c, curr_desc, curr_t, curr_c, prev_rc))

    for vehicule, occurrences in par_vehicule.items():
        total = len(occurrences)
        for curr_rc, prev_desc, prev_t, prev_c, curr_desc, curr_t, curr_c, prev_rc in occurrences:
            anomalies.append(Anomaly(
                "Cohérence", "Compteur décroissant", curr_rc, vehicule,
                "Le compteur diminue par rapport à l'entrée précédente pour ce véhicule",
                f"{prev_desc} le {prev_t} : {prev_c} -> {curr_desc} le {curr_t} : {curr_c} "
                f"(réponse précédente : {prev_rc}) -- écart : {prev_c - curr_c:.0f} km -- "
                f"{total} occurrence(s) de compteur décroissant au total pour ce véhicule",
            ))
    return anomalies


def check_coherence_nom_conducteur(rows):
    """Signale les paires de noms de conducteur suffisamment proches pour être une variante
    probable (faute de frappe) plutôt que deux personnes différentes -- à vérifier, pas une
    anomalie certaine (voir SIMILARITE_NOM_SEUIL)."""
    anomalies = []
    noms = sorted({(r.get("Nom du conducteur") or "").strip() for r in rows if (r.get("Nom du conducteur") or "").strip()})

    # Une seule entrée par paire proche détectée, pour ne pas noyer le log de doublons de signalement
    signalees = set()
    for i, nom1 in enumerate(noms):
        for nom2 in noms[i + 1:]:
            ratio = difflib.SequenceMatcher(None, nom1.lower(), nom2.lower()).ratio()
            if SIMILARITE_NOM_SEUIL <= ratio < 1.0:
                pair_key = tuple(sorted((nom1, nom2)))
                if pair_key in signalees:
                    continue
                signalees.add(pair_key)
                # Rattache l'anomalie à la première réponse trouvée pour nom2 (le plus rare des deux,
                # généralement la variante), pour donner un point d'entrée dans les données.
                rc = next((r.get("Response Code", "") for r in rows if (r.get("Nom du conducteur") or "").strip() == nom2), "")
                anomalies.append(Anomaly(
                    "Cohérence", "Variante nom conducteur", rc, "",
                    "Deux noms de conducteur proches -- possible faute de frappe/variante",
                    f"'{nom1}' vs '{nom2}' (similarité {ratio:.2f})",
                ))
    return anomalies


def check_coherence_distance_gps(rows):
    """Deux vérifications distinctes, plus fiables qu'une simple comparaison compteur/GPS
    (voir échange avec Lanja -- les appareils utilisés ont une précision GPS très variable,
    jusqu'à ~16 km sur certains relevés réels, donc un seuil fixe produit trop de faux positifs) :

    1. La distance GPS (vol d'oiseau) ne peut géométriquement pas dépasser la distance
       réellement parcourue (compteur) -- sauf marge de précision des appareils. On utilise la
       précision (accuracy, en mètres) réellement rapportée par chaque point GPS de la ligne
       comme tolérance dynamique, plutôt qu'un seuil arbitraire unique. Si ça arrive quand même
       largement au-delà de cette tolérance, la capture GPS (pas le compteur) est en cause.
    2. Le compteur d'un seul trajet doit rester dans une plage plausible dans l'absolu, quel
       que soit le GPS -- attrape les valeurs corrompues (ex. plusieurs millions de km vus dans
       les données réelles) que le test de distance négative ne détecte pas s'il s'agit d'une
       valeur positive mais absurde.
    """
    anomalies = []
    for r in rows:
        if (r.get("Type d'intervention") or "").strip() != TYPE_TRAJET:
            continue
        rc = r.get("Response Code", "")
        vehicule = r.get("Véhicule > Immatriculaton", "")
        dist_compteur = parse_float(r.get("Distance parcourue (en Km)"))

        if dist_compteur is not None and abs(dist_compteur) > COMPTEUR_TRAJET_MAX_KM:
            anomalies.append(Anomaly(
                "Cohérence", "Compteur trajet hors plage plausible", rc, vehicule,
                f"Distance compteur d'un seul trajet > {COMPTEUR_TRAJET_MAX_KM} km -- valeur probablement corrompue",
                f"Distance parcourue (en Km): {dist_compteur}",
            ))

        p1 = parse_latlng(r.get("Point de départ"))
        p2 = parse_latlng(r.get("Point d'arrivée"))
        if dist_compteur is None or not p1 or not p2:
            continue
        dist_gps = haversine_km(p1, p2)

        acc_depart_m = parse_float(r.get("Point de départ (accuracy)")) or GPS_ACCURACY_DEFAUT_M
        acc_arrivee_m = parse_float(r.get("Point d'arrivée (accuracy)")) or GPS_ACCURACY_DEFAUT_M
        tolerance_km = max(GPS_TOLERANCE_PLANCHER_KM, (acc_depart_m + acc_arrivee_m) / 1000.0 * GPS_TOLERANCE_MARGE)

        if dist_gps > dist_compteur + tolerance_km:
            anomalies.append(Anomaly(
                "Cohérence", "Distance GPS supérieure au compteur", rc, vehicule,
                "Distance GPS (vol d'oiseau) supérieure à la distance compteur au-delà de la "
                "marge de précision des appareils -- géométriquement impossible, capture GPS suspecte",
                f"Compteur: {dist_compteur} km / GPS: {dist_gps:.1f} km / "
                f"Tolérance (précision appareils): {tolerance_km:.1f} km "
                f"(précision départ: {acc_depart_m:.0f} m, arrivée: {acc_arrivee_m:.0f} m)",
            ))
    return anomalies


def check_coherence(rows):
    return (
        check_coherence_compteur_croissant(rows)
        + check_coherence_nom_conducteur(rows)
        + check_coherence_distance_gps(rows)
    )


# ---------------------------------------------------------------------------
# Rapport Excel — log unique avec suivi Nouveau / Toujours ouvert / Résolu
# (mêmes principes que verify_maintenance_preventive.py -- voir sa docstring
# anomaly_key pour l'incident du 29/07/2026 qui a motivé d'inclure un identifiant
# d'entité dans la clé, ici Véhicule)
# ---------------------------------------------------------------------------

def anomaly_key(a):
    return (a.dimension, a.subdimension, a.response_code, a.description, a.vehicule)


def find_child_item_id(token, drive_id, folder_item_id, file_name):
    resp = requests.get(
        f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{folder_item_id}/children",
        headers={"Authorization": f"Bearer {token}"},
        timeout=60,
    )
    raise_for_status_verbose(resp)
    for item in resp.json().get("value", []):
        if item.get("name") == file_name:
            return item.get("id")
    return None


def download_existing_log(token, drive_id, folder_item_id, file_name):
    item_id = find_child_item_id(token, drive_id, folder_item_id, file_name)
    if not item_id:
        return []

    resp = requests.get(
        f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item_id}/content",
        headers={"Authorization": f"Bearer {token}"},
        timeout=60,
    )
    raise_for_status_verbose(resp)

    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb["Anomalies"]
    headers = [c.value for c in ws[1]]
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        if row.get("Dimension"):
            rows.append(row)
    return rows


def build_submitted_on_lookup(rows):
    lookup = {}
    for r in rows:
        rc = r.get("Response Code")
        dt = parse_dt(r.get("Submitted On"))
        if rc and dt:
            lookup[rc] = dt.strftime("%d/%m/%Y")
    return lookup


def merge_with_log(current_anomalies, existing_rows, today_str, submitted_on_lookup=None,
                    existing_response_codes=None):
    """Voir merge_with_log dans verify_maintenance_preventive.py pour le détail des règles
    (Nouveau / Toujours ouvert / Résolu / Supprimé) -- logique identique."""
    submitted_on_lookup = submitted_on_lookup or {}
    existing_response_codes = existing_response_codes or set()
    existing_open = {}
    already_closed = []
    for row in existing_rows:
        key = (row.get("Dimension"), row.get("Sous-dimension"), row.get("Response Code"),
               row.get("Description"), row.get("Véhicule"))
        if row.get("Statut") in ("Résolu", "Supprimé"):
            already_closed.append(row)
        else:
            existing_open[key] = row

    merged = []
    seen_keys = set()
    new_count = 0

    for a in current_anomalies:
        key = anomaly_key(a)
        seen_keys.add(key)
        existing = existing_open.get(key)
        if not existing:
            new_count += 1
        merged.append({
            "Dimension": a.dimension,
            "Sous-dimension": a.subdimension,
            "Response Code": a.response_code,
            "Véhicule": a.vehicule,
            "Description": a.description,
            "Détails": a.details,
            "Statut": "Toujours ouvert" if existing else "Nouveau",
            "Première détection": existing.get("Première détection") if existing else today_str,
            "Dernière détection": today_str,
            "Date de résolution": "",
        })

    resolved_count = 0
    deleted_count = 0
    for key, row in existing_open.items():
        if key not in seen_keys:
            row = dict(row)
            if row.get("Response Code") not in existing_response_codes:
                row["Statut"] = "Supprimé"
                row["Date de résolution"] = row.get("Première détection", "")
                deleted_count += 1
            else:
                row["Statut"] = "Résolu"
                row["Date de résolution"] = submitted_on_lookup.get(row.get("Response Code"), today_str)
                resolved_count += 1
            merged.append(row)

    merged.extend(already_closed)
    return merged, new_count, resolved_count, deleted_count


def build_log_report(merged_rows, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Anomalies"
    ws.append(LOG_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in merged_rows:
        ws.append([row.get(h, "") for h in LOG_HEADERS])

    open_rows = [r for r in merged_rows if r.get("Statut") not in ("Résolu", "Supprimé")]
    summary = wb.create_sheet("Résumé")
    summary.append(["Dimension", "Nombre d'anomalies ouvertes"])
    for cell in summary[1]:
        cell.font = Font(bold=True)
    counts = Counter(r.get("Dimension") for r in open_rows)
    for dimension in DIMENSIONS_AUTOMATISEES:
        summary.append([dimension, counts.get(dimension, 0)])
    summary.append(["Fiabilité", "N/A — contrôle documentaire manuel, non automatisé"])
    summary.append(["Total ouvert", len(open_rows)])

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Microsoft Graph : upload SharePoint + email
# (identique à verify_maintenance_preventive.py)
# ---------------------------------------------------------------------------

def graph_token(tenant_id, client_id, client_secret):
    resp = requests.post(
        f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token",
        data={
            "grant_type": "client_credentials",
            "client_id": client_id,
            "client_secret": client_secret,
            "scope": "https://graph.microsoft.com/.default",
        },
        timeout=30,
    )
    raise_for_status_verbose(resp)
    return resp.json()["access_token"]


def resolve_share_link(token, share_url):
    b64 = base64.b64encode(share_url.encode("utf-8")).decode("utf-8")
    encoded = "u!" + b64.replace("+", "-").replace("/", "_").rstrip("=")
    resp = requests.get(
        f"https://graph.microsoft.com/v1.0/shares/{encoded}/driveItem",
        headers={"Authorization": f"Bearer {token}"},
        timeout=30,
    )
    raise_for_status_verbose(resp)
    data = resp.json()
    return data["parentReference"]["driveId"], data["id"]


def upload_to_sharepoint(token, drive_id, folder_item_id, file_path, file_name):
    with open(file_path, "rb") as f:
        content = f.read()

    item_id = find_child_item_id(token, drive_id, folder_item_id, file_name)
    if not item_id:
        resp = requests.post(
            f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{folder_item_id}/children",
            headers={"Authorization": f"Bearer {token}"},
            json={"name": file_name, "file": {}, "@microsoft.graph.conflictBehavior": "replace"},
            timeout=60,
        )
        raise_for_status_verbose(resp)
        item_id = resp.json()["id"]

    resp = requests.put(
        f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item_id}/content",
        headers={"Authorization": f"Bearer {token}", "Content-Type": "application/octet-stream"},
        data=content,
        timeout=120,
    )
    raise_for_status_verbose(resp)
    return resp.json()


def send_confirmation_email(token, sender, recipients, file_name, open_count, new_count,
                             resolved_count, deleted_count, counts_by_dimension):
    lines = "".join(f"<li>{dim} : {n}</li>" for dim, n in counts_by_dimension.items())
    body_html = (
        f"<p>Vérification de données \"Carnet de bord\" exécutée.</p>"
        f"<p>Anomalies actuellement ouvertes : <b>{open_count}</b> "
        f"(dont {new_count} nouvelles) — {resolved_count} résolue(s) et {deleted_count} "
        f"supprimée(s) (réponse mWater disparue) depuis la dernière exécution.</p>"
        f"<ul>{lines}</ul>"
        f"<p>Fiabilité n'est pas incluse (contrôle documentaire manuel).</p>"
        f"<p>Log mis à jour sur SharePoint : <b>{file_name}</b></p>"
    )
    resp = requests.post(
        f"https://graph.microsoft.com/v1.0/users/{sender}/sendMail",
        headers={"Authorization": f"Bearer {token}"},
        json={
            "message": {
                "subject": f"Vérification de données — {file_name}",
                "body": {"contentType": "HTML", "content": body_html},
                "toRecipients": [{"emailAddress": {"address": addr.strip()}} for addr in recipients.split(",")],
            }
        },
        timeout=30,
    )
    raise_for_status_verbose(resp)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    mwater_username = os.environ["MWATER_USERNAME"].strip()
    mwater_password = os.environ["MWATER_PASSWORD"].strip()

    azure_tenant_id = os.environ["AZURE_TENANT_ID"].strip()
    azure_client_id = os.environ["AZURE_CLIENT_ID"].strip()
    azure_client_secret = os.environ["AZURE_CLIENT_SECRET"].strip()

    sharepoint_folder_link = os.environ["SHAREPOINT_FOLDER_LINK"].strip()

    email_sender = os.environ["EMAIL_SENDER"].strip()
    email_recipients = os.environ["EMAIL_RECIPIENTS"].strip()

    print("Authentification mWater...")
    client_id = mwater_login(mwater_username, mwater_password)

    print("Téléchargement du datagrid Carnet de bord...")
    rows = download_datagrid(DATAGRID_CARNET, client_id)
    print(f"  {len(rows)} lignes")

    print("Application des règles de vérification (5 dimensions automatisées)...")
    anomalies = []
    anomalies += check_completude(rows)
    anomalies += check_promptitude(rows)
    anomalies += check_validite(rows)
    anomalies += check_unicite(rows)
    anomalies += check_coherence(rows)
    print(f"  {len(anomalies)} anomalies détectées")

    print("Authentification Microsoft Graph...")
    token = graph_token(azure_tenant_id, azure_client_id, azure_client_secret)

    print("Résolution du lien SharePoint...")
    sharepoint_drive_id, sharepoint_folder_item_id = resolve_share_link(token, sharepoint_folder_link)

    print("Téléchargement du log existant sur SharePoint...")
    existing_rows = download_existing_log(token, sharepoint_drive_id, sharepoint_folder_item_id, LOG_FILE_NAME)
    print(f"  {len(existing_rows)} lignes déjà présentes dans le log")

    today_str = datetime.now().strftime("%d/%m/%Y")
    submitted_on_lookup = build_submitted_on_lookup(rows)
    existing_response_codes = {r.get("Response Code", "") for r in rows}
    merged_rows, new_count, resolved_count, deleted_count = merge_with_log(
        anomalies, existing_rows, today_str,
        submitted_on_lookup=submitted_on_lookup,
        existing_response_codes=existing_response_codes)
    open_rows = [r for r in merged_rows if r.get("Statut") not in ("Résolu", "Supprimé")]
    print(f"  {len(open_rows)} anomalies ouvertes ({new_count} nouvelles, {resolved_count} résolues, "
          f"{deleted_count} supprimées)")

    output_path = f"/tmp/{LOG_FILE_NAME}"
    build_log_report(merged_rows, output_path)
    print(f"Rapport généré : {output_path}")

    print("Dépôt du log mis à jour sur SharePoint...")
    upload_to_sharepoint(token, sharepoint_drive_id, sharepoint_folder_item_id, output_path, LOG_FILE_NAME)

    print("Envoi de l'email de confirmation...")
    counts_by_dimension = Counter(r.get("Dimension") for r in open_rows)
    send_confirmation_email(token, email_sender, email_recipients, LOG_FILE_NAME,
                             len(open_rows), new_count, resolved_count, deleted_count, counts_by_dimension)

    print("Terminé.")


if __name__ == "__main__":
    main()
